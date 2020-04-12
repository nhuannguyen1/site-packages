import pandas as pd
import openpyxl 
from pynvn.path.ppath import PathSteel
class hdata:
    def __init__(self,pathfile = None,
                sheetname = "AZB-10",
                rowindexstart = 41,
                columnmh = 3,
                columntotp = 18
                ):
                self.pathfile = pathfile
                self.sheetname = sheetname
                self.rowindexstart = rowindexstart
                self.columnmh = columnmh
                self.columntotp = columntotp

    def hldatakid(self):

        # workbook object is created
        self.wb_obj = openpyxl.load_workbook(self.pathfile,
                                                    data_only=True) 
                
        # get sheet name file 
        sheet = self.wb_obj.get_sheet_by_name(self.sheetname)
        #self.wb_obj.save(self.pathfile)
        # find max row
        self.mr = sheet.max_row
        i = 0 
        sum = 0
        cli = 41
        while cli < self.mr:
            valcell = sheet.cell(row=cli, column=self.columnmh).value
            if valcell == None:
                for ci in range (cli,self.mr):
                    vaulenone = sheet.cell(row=ci, column=self.columnmh).value
                    if  vaulenone == None:
                        value = sheet.cell(row=ci, column=self.columntotp).value
                        valuem = value if value is not None else 0 
                        sum += valuem
                        i += 1
                    else:
                        sheet.cell(row=cli - 1, column=self.columntotp).value = sum
                        cli = cli + i
                        i = 0
                        sum = 0
                        break             
            cli += 1 
        self.wb_obj.save(self.pathfile)
    def getdatafromkid(self):
        

        
