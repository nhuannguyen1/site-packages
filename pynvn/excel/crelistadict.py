from openpyxl import Workbook
import openpyxl
import xlwings as xw 
from pynvn.excel import mrowandmcolum
from pynvn.excel import (convertrangaphatonunber,
                            returnrangelastcolumn)
class credict:
    """ create dict for value and key """
    def __init__ (self, ws = None,
                        rangea = "C7:C1000", 
                        pathfull = None, 
                        namesheet ="PTVT1", 
                        engine ="openpyxl"):
        self.ws = ws
        self.rangea = rangea
        self.engine = engine
        if engine =="openpyxl":
            wb = openpyxl.load_workbook(pathfull, 
                                        read_only=True,
                                        data_only= True
                                        )
            self.ws = wb[namesheet]
        else:
            wb = xw.Book(pathfull)
            self.ws = wb.sheets[namesheet]
            self.lastrow = self.ws.range('A' + str(wb.sheets[0].cells.last_cell.row)).end('up').row
            self.rangea = returnrangelastcolumn(stringrang=rangea,lrow=self.lastrow)
    def reindexrownotnone(self):
        """ renturn index which value not none"""
        if self.engine =="openpyxl":
            key_list = [cell.row for row in self.ws[self.rangea]  for cell in row if cell.value != None]
        else:
            key_list = [cell.row for cell in self.ws.range(self.rangea) if cell.value != None]
        return key_list

    def revaluerownotnone(self):
        """ renturn value which value not none"""
        if self.engine =="openpyxl":
            value_list = [cell.value for row in self.ws[self.rangea] for cell in row if cell.value != None ]
            
        else:
            value_list = [cell.value for cell in self.ws.range(self.rangea) if cell.value != None]

        return value_list
    def redictvaluesandvaluecol(self, columnumber = 4 ):
        """ return dict value and value column"""
        arrch = []
        res = list(zip(self.reindexrownotnone(), self.reindexrownotnone()[1:] + self.reindexrownotnone()[:1])) 
        for eler in res:
            arrchild = self.__listchild(eler,columnumber=columnumber)
            arrch.append(arrchild)
            arrchild = None
        dictionary = dict(zip(self.revaluerownotnone(), arrch))
        return dictionary
    
    def returndictvaluebyindexcolumnandrow(self, 
                                            value_criteria_range, 
                                            range_col ="D7:D1000", 
                                            indexcolumn = [5,6,8]):
        """ return dict by value rangce_col and indexcolumn """
        if self.engine =="openpyxl":
                indexrcevalu = [[self.valuebycol_row(cell.row,indexcolumn[0]),

                                self.valuebycol_row(cell.row,indexcolumn[1]),

                                self.valuebycol_row(cell.row,indexcolumn[2]),

                                ] for range_cell in  self.ws[range_col]\

                                for cell in range_cell  if  cell.value == value_criteria_range

                                ]
                                
        else:
                indexrcevalu = [[self.valuebycol_row(cell.row,indexcolumn[0]),
                                self.valuebycol_row(cell.row,indexcolumn[1]),
                                self.valuebycol_row(cell.row,indexcolumn[2])] \
                                for cell in self.ws.range(range_col) if cell.value == value_criteria_range]

        return indexrcevalu
            
    def valuebycol_row(self,irow,icolumn):
        """ return value cell by index column and row"""
        if self.engine =="openpyxl":
            valuebycolr = self.ws.cell(row=irow,column = icolumn).value 
        else:
            valuebycolr = self.ws.range(irow,icolumn).value 
        return valuebycolr

    def __listchild(self,ele,columnumber = 4):
        s,t = ele
        if self.engine =="openpyxl":
            arrchild = [self.ws.cell(row=ie,
                        column = columnumber).value for ie in range(s,t) if self.ws.cell(row=ie,
                                                                                        column = columnumber).value != None]
        else:
            arrchild = [[self.ws.range((ie,columnumber)).row,
                        self.ws.range((ie,columnumber)).value] for ie in range(s,t) if self.ws.range((ie,
                                                                                                    columnumber)).value != None] 
        return arrchild

    def listothercell (self,irow,icolumn):
        """ return value of column sheet ptvl1"""
        valuebycolr = self.ws.range(irow,icolumn).value 
        return valuebycolr
