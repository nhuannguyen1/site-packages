import openpyxl as xl
from tkinter import messagebox
from pynvn.path.ppath import refullpath
from pynvn.excel import returnsheet,repathlinkexcel,relistsheet,colnum_string

class hexcel:
    """copy excel to excel"""
    def __init__(self,wsheet = None,
                    dpath = None,
                    namesheet = "TONG HOP HM",
                    namefile = None,
                    keyhm = "HM."
                ):
                self.dpath = dpath
                self.wsheet = wsheet
                self.namesheet = namesheet
                self.namefile = namefile
                self.keyhm = keyhm
                self.fpath = refullpath(dirpath=self.dpath,
                                        filename=self.namefile)
                print ("self.fpath",self.fpath )

    def habz30 (self):
        """handling data azb-30 sheet"""
        # return name sheet for link fomular
        #pfile = repathlinkexcel(dpath=self.dpath,namefile=self.namefile,namesheet=self.namesheet )

        #=SUMPRODUCT(('HM.HÀNG RÀO'!$A$38:$A$61='AZB-30'!C16)*('HM.HÀNG RÀO'!$B$38:$B$61 = 'AZB-30'!J3)*'HM.HÀNG RÀO'!$I$38:$I$61)
        # return list of list 
        lsheetnames = relistsheet(self.fpath)
        """
        for lsheet in lsheetnames:
            if self.keyhm in lsheet:
                pfile = repathlinkexcel(dpath=self.dpath,
                                        namefile=self.namefile,
                                        namesheet=lsheet)
        """
        stt = "'" + "AZB-30" + "'"
        for k in range (10,30):
            cln = colnum_string(k)
            hmname =  self.wsheet.cell(row=3, column=k).value
            #print ("hmname,lsheetnames",hmname,lsheetnames)
            if hmname in lsheetnames:
                print ("hmname,lsheetnames",hmname,lsheetnames)
                pfile = repathlinkexcel(dpath=self.dpath,
                                        namefile=self.namefile,
                                        namesheet=hmname)
                for i in range(14,474):
                    valuene = '=SUMPRODUCT(({0}!$A$38:$A$61={1}!C{2})*({0}!$B$38:$B$61 = "{3}")*{0}!$I$38:$I$61)'.format(pfile,stt,i,hmname)

                    self.wsheet.cell(row=i, column=k).value = valuene



                
