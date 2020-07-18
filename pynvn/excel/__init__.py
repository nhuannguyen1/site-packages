from pynvn.excel.toexcel import toexcel
from tkinter import messagebox
import openpyxl as xl
import string
from pynvn.string  import sepnumberandstrfromstr,returnrangewolastrow
import xlwings as xw
from xlwings.constants import DeleteShiftDirection
#from pynvn.excel.exception import closebookactive
def returnsheet (path, namesheet = "TONG HOP HM"):
    """ return sheet name by index and path excel """
    wb1 = xl.load_workbook(filename=path)
    ws1 = wb1[namesheet]
    return ws1
def  repathlinkexcel (dpath,namefile,namesheet):
    """ return path link excel """
    pfile = "'" + dpath + "/" + "[" + namefile + "]" + namesheet + "'"
    return pfile

def colnum_string(n):
    """conver colum number become string"""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def relistsheet(path):
    """ return all sheet of excel file """
    wb1 = xl.load_workbook(filename=path)
    names = wb1.sheetnames
    return names
def returnactivesheet(path):
    """ return active sheet """
    wb1 = xl.load_workbook(filename=path)
    sheet = wb1.active
    return sheet
def returnsheetbyname(path = None, sheetname = "PTVT"):
    """return sheet by name """
    wb1 = xl.load_workbook(filename=path)
    return wb1[sheetname]
def mrowandmcolum (path = None):
    "return maxcolumn and max row column"
    wb1 = xl.load_workbook(filename=path)
    wsactive = wb1.active
    return [wsactive.max_row,wsactive.max_column]

def col2num(col):
    """Return number corresponding to excel-style column."""
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num
def convertrangaphatonunber (rangapha = None ):
    """ convert rang anphab to cell number"""
    a,n,s,a1,n1 =  sepnumberandstrfromstr(sstr=rangapha)
    return  (col2num(a),int(n)),(col2num(a1),int(n1))
def returnrangelastcolumn(stringrang,lrow = 100):
    """ 
        return range excel by range and by last row
        ex: 'A5:A100' ---> 'A5:2000'
        2000 is last row replace in stringrang
    """
    return returnrangewolastrow(sstr=stringrang) + str(lrow)

def delrowbyindexcell (incolumndel = "C", 
                        valueofindexcoldel = None, 
                        wb = None,
                        namesheet = None,
                        startrow =1,
                        endrow = 1000,
                        valuetoendrow = "VTC"
                        ):
    """ delete row by value of cell """
    for i in range (startrow,
                        endrow):
        valuecompare =wb.sheets[namesheet].range(i,
                                                incolumndel ).value 
        k = i
        if (valuecompare == None or valuecompare == ""):
            while True:
                wb.sheets[namesheet].range('{0}:{0}'.format(k)).api.Delete(DeleteShiftDirection.xlShiftUp)
                if (wb.sheets[namesheet].range(k,incolumndel).value != None and (wb.sheets[namesheet].range(k,incolumndel).value != "")) :
                    break
        if wb.sheets[namesheet].range(k,incolumndel).value == valuetoendrow :
            break
def cellcoordbyvalue(max_row = 20, 
                    min_row = 1 , 
                    max_col = 10,
                    min_col= 0 , 
                    sheet = None,
                    valuetofile = None
                    ):
    """find a cell coordinate containing a value """
    if type(min_col) == str:
        min_col = col2num(min_col)
    if type(max_col) == str:
        max_col = col2num(max_col)
    
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if sheet.range((row,col)).value == valuetofile:
                #print("The Row is: "+str(row)+" and the column is "+str(col))
                return [row,col]
                break
    else:
        messagebox.showerror ("Error", "can not find name value {0} at sheet {1} location column index {2} to {3}, check again".format(valuetofile,
                                                                                                                                        sheet.name,
                                                                                                                                        min_col,
                                                                                                                                        max_col))
        raise ValueError("can not find name value")
        #xw.App(visible=False).quit()
        #raise closebookactive()

    #lrowcol = [row for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1) if sheet.range((row,col)).value == valuetofile]
    return lrowcol
def lcellindexbyvalue(lvalue, 
                        max_row = 20, 
                        min_row = 1 , 
                        max_col = 10,
                        min_col= 0 , 
                        sheet = None,
                    ):
    """ return list cell index row by list value """
    return  [cellcoordbyvalue(max_col=max_col,
                                        max_row=max_row,
                                        min_row=min_row,
                                        min_col=min_col,
                                        sheet=sheet,
                                        valuetofile=value)[0] for value in lvalue
            ]
def openexcelbyxl (pathex):
    """ open excel by pathex"""
    xw.Book(pathex)
def valuebyindexrowcell(lindexcell = None,
                        col = None,
                        sheet = None):
    """ return list value with indexcell """
    return [sheet.range("{0}{1}".format(col,indexcell)).value for indexcell in lindexcell]
def closeallfileexcel (namek_ofpname):
    """ close all file excel by list if it is openning filter by namek_ofpname """
    listwb = xw.books
    for wb in listwb:
        if namek_ofpname in wb.name:
            res = messagebox.askyesno("Check file is opening","file name {0} is opening, Please close it, do you close and save it ?".format(wb.name))
            if res:
                quit_excel(wb)
            else:
                break
def quit_excel(wb):
    """
    quit workbook 
    wb: workbook object from xlwings
    """
    # look if PERSONAL.XLSB is in the list of books associated with the Excel App
    if "PERSONAL.XLSB" in [b.name for b in wb.app.books]:
        if len(wb.app.books) == 2: 
            wb.app.quit()
        else:
            wb.close()
    else:
        if len(wb.app.books) == 1: 
            wb.app.quit()
        else:
            wb.close()
def rangecopyrefsamesheet (sheet = None,
                            formulasfirstcell = None,
                            col_index = None,
                            startrow = None,
                            endrow = None):
        """copy range have reference at col_index"""

        sheet.range(startrow,col_index).value =   formulasfirstcell 

        vtformulas = sheet.range(startrow,
                                        col_index).formula

        sheet.range("{0}{1}:{0}{2}".format(colnum_string(col_index),
                                            startrow,
                                            endrow)).formula = vtformulas

def returnvaluekeyim (cola,listvalue_im,sheet,indexrow_im):
        """ return value by cell index and list value  """
        for count,numberint in enumerate(indexrow_im,0):
            sheet.range("{0}{1}".format(cola,
                                        numberint)).value = listvalue_im[count]
def activesheet():
    """ return active sheet of excel file """

    try: 
        self.__sheetdesactive = xw.sheets.active
    except:
        messagebox.showerror("Error","No excel file has been opened yet, please open the excel file")